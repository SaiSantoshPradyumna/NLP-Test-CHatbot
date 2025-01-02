
import streamlit as st
from huggingface_hub import InferenceClient
import PyPDF2
# import csv
asd
from openpyxl import load_workbook

client = InferenceClient(
    model="microsoft/Phi-3.5-mini-instruct",
    token="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"
)

def extract_text_from_pdf(file):
    text = ""
    file.seek(0)  # Ensure the pointer is at the start of the file
    reader = PyPDF2.PdfReader(file)
    for page_num in range(len(reader.pages)):
        page = reader.pages[page_num]
        text += page.extract_text() or ""  # Extract text from each page
    return text

def extract_text_from_csv(file):
    text = ""
    file.seek(0)  # Move to the start of the file
    reader = csv.reader(file.read().decode('utf-8').splitlines())
    for row in reader:
        text += " ".join(row) + "\n"
    return text


def extract_text_from_xlsx(file):
    text = ""
    file.seek(0)  # Ensure the pointer is at the start of the file
    workbook = load_workbook(filename=file, data_only=True)  # Read directly from file-like object
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        text += f"\nSheet: {sheet}\n"
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
    return text

def truncate_text(text, max_length=6000):
    if len(text) > max_length:
        return text[:max_length] + "\n... [Truncated]"
    return text

def ask_question_based_on_document(context_messages, document_text, question):
    context_messages.append({"role": "user", "content": question})
    context = (
        f"Based on the provided document, answer the following question.\n\n"
        f"Document Content:\n{truncate_text(document_text)}\n\n"
        f"Conversation history:\n"
        f"{' '.join([f'{msg['role']}: {msg['content']}' for msg in context_messages])}\n\n"
    )
    response = client.chat_completion(
        messages=[{"role": "user", "content": context}],
        max_tokens=300,
        stream=False,
    )
    context_messages.append({"role": "assistant", "content": response.choices[0].message.content})
    return response.choices[0].message.content

def ask_question_without_document(context_messages, question):
    context_messages.append({"role": "user", "content": question})
    context = (
        f"Conversation history:\n"
        f"{' '.join([f'{msg['role']}: {msg['content']}' for msg in context_messages])}\n\n"
        f"Answer the user's next question."
    )
    response = client.chat_completion(
        messages=[{"role": "user", "content": context}],
        max_tokens=300,
        stream=False,
    )
    context_messages.append({"role": "assistant", "content": response.choices[0].message.content})
    return response.choices[0].message.content

st.title("Document-based Q&A Chatbot")
st.write("Upload a document and ask questions, or just ask questions directly.")

context_messages = []
document_text = None

if st.sidebar.radio("Choose an option", ("Upload Document", "Ask Without Document")) == "Upload Document":
    uploaded_file = st.file_uploader("Upload a PDF, CSV, or XLSX file", type=["pdf", "csv", "xlsx"])
    
    if uploaded_file is not None:
        if uploaded_file.name.endswith(".pdf"):
            document_text = extract_text_from_pdf(uploaded_file)
        elif uploaded_file.name.endswith(".csv"):
            document_text = extract_text_from_csv(uploaded_file)
        elif uploaded_file.name.endswith(".xlsx"):
            document_text = extract_text_from_xlsx(uploaded_file)
        
        st.success("Document loaded. You can now ask questions about it.")

st.write("### Ask a question:")
user_question = st.text_input("Your question:")

if st.button("Get Answer"):
    if document_text:
        answer = ask_question_based_on_document(context_messages, document_text, user_question)
    else:
        answer = ask_question_without_document(context_messages, user_question)
    st.write("**Model:**", answer)

